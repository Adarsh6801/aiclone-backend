"""
AI Clone Backend — FastAPI + NVIDIA NIM API (Llama 4)
Run: uvicorn main:app --reload --port 8000

All personal training data is in profile.py — edit that file to train your clone.
"""
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional
import requests, os, json, uuid, random, sqlite3, shutil, hashlib, secrets
from dotenv import load_dotenv

# Import your personal training data from profile.py
from profile import CLONE_SYSTEM_PROMPT, HIRE_SYSTEM_PROMPT

load_dotenv()

app = FastAPI(title="AI Clone API", version="4.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
    "http://localhost:5173",
    "http://13.61.34.71:8000",
    "https://devadarsh.xyz",
    "https://www.devadarsh.xyz",
    "https://api.devadarsh.xyz",   # if needed
    "https://your-netlify-site.netlify.app",  # keep for now
],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── UPLOADS & DATABASE SETUP ───────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(__file__)
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
DB_PATH     = os.path.join(BASE_DIR, "ayclone.db")
os.makedirs(UPLOADS_DIR, exist_ok=True)

# Serve uploaded images as static files at /uploads/filename
app.mount("/uploads", StaticFiles(directory=UPLOADS_DIR), name="uploads")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cur  = conn.cursor()
    # Referrals table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS referrals (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT    NOT NULL,
            email      TEXT,
            role       TEXT,
            company    TEXT,
            quote      TEXT    NOT NULL,
            photo_path TEXT,
            relation   TEXT,
            verified   INTEGER DEFAULT 0,
            created_at TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    # Feedback table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS feedback (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            message    TEXT    NOT NULL,
            rating     INTEGER DEFAULT 5,
            created_at TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    # Blog admin table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS blog_admins (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT    NOT NULL UNIQUE,
            password   TEXT    NOT NULL,
            created_at TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    # Blogs table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS blogs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            title       TEXT    NOT NULL,
            description TEXT    NOT NULL,
            category    TEXT    DEFAULT 'Tech',
            image_url   TEXT    DEFAULT '',
            likes       INTEGER DEFAULT 0,
            created_at  TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    # Blog comments table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS blog_comments (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            blog_id    INTEGER NOT NULL,
            name       TEXT    DEFAULT 'Anonymous',
            body       TEXT    NOT NULL,
            created_at TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (blog_id) REFERENCES blogs(id) ON DELETE CASCADE
        )
    """)
    # Blog likes table (track by ip fingerprint)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS blog_likes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            blog_id    INTEGER NOT NULL,
            fingerprint TEXT   NOT NULL,
            created_at TEXT    DEFAULT (datetime('now','localtime')),
            UNIQUE(blog_id, fingerprint)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS learning_tracks (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            title       TEXT    NOT NULL,
            description TEXT    DEFAULT '',
            kind        TEXT    NOT NULL DEFAULT 'tree',
            priority    INTEGER DEFAULT 0,
            color       TEXT    DEFAULT '',
            tags_json   TEXT    DEFAULT '[]',
            is_news     INTEGER DEFAULT 0,
            created_at  TEXT    DEFAULT (datetime('now','localtime')),
            updated_at  TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS learning_topics (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            track_id    INTEGER NOT NULL,
            parent_id   INTEGER,
            title       TEXT    NOT NULL,
            description TEXT    DEFAULT '',
            priority    INTEGER DEFAULT 0,
            links_json  TEXT    DEFAULT '[]',
            tags_json   TEXT    DEFAULT '[]',
            is_news     INTEGER DEFAULT 0,
            created_at  TEXT    DEFAULT (datetime('now','localtime')),
            updated_at  TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (track_id) REFERENCES learning_tracks(id) ON DELETE CASCADE,
            FOREIGN KEY (parent_id) REFERENCES learning_topics(id) ON DELETE CASCADE
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS learning_progress (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            track_id    INTEGER NOT NULL,
            topic_id    INTEGER NOT NULL,
            fingerprint TEXT    NOT NULL,
            status      TEXT    NOT NULL DEFAULT 'pending',
            updated_at  TEXT    DEFAULT (datetime('now','localtime')),
            UNIQUE(topic_id, fingerprint)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS learning_comments (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            track_id    INTEGER NOT NULL,
            parent_id   INTEGER,
            name        TEXT    NOT NULL,
            body        TEXT    NOT NULL,
            fingerprint TEXT    DEFAULT '',
            is_author   INTEGER DEFAULT 0,
            created_at  TEXT    DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (parent_id) REFERENCES learning_comments(id) ON DELETE CASCADE
        )
    """)
    # Insert default admin if not exists (username: admin, password: admin123)
    default_pw = hashlib.sha256("admin123".encode()).hexdigest()
    cur.execute("INSERT OR IGNORE INTO blog_admins (username, password) VALUES (?, ?)", ("admin", default_pw))
    conn.commit()
    conn.close()

init_db()


# ── NVIDIA NIM CONFIG ──────────────────────────────────────────────────────────
NVIDIA_API_KEY = os.getenv("NVIDIA_API_KEY", "YOUR_NVIDIA_API_KEY_HERE")
NVIDIA_URL     = "https://integrate.api.nvidia.com/v1/chat/completions"
MODEL          = "meta/llama-4-maverick-17b-128e-instruct"
# Other models you can try in profile.py or here:
# "meta/llama-3.3-70b-instruct"      ← more accurate, slightly slower
# "mistralai/mistral-large-2-instruct"


def nvidia_call(messages: list, temperature: float = 0.8, max_tokens: int = 500) -> str:
    """Call NVIDIA NIM API and return the text reply."""
    headers = {
        "Authorization": f"Bearer {NVIDIA_API_KEY}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    payload = {
        "model": MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
        "top_p": 1.0,
        "stream": False,
    }
    resp = requests.post(NVIDIA_URL, headers=headers, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()


def ask(system: str, user: str, temperature: float = 0.8, max_tokens: int = 500) -> str:
    return nvidia_call(
        [{"role": "system", "content": system}, {"role": "user", "content": user}],
        temperature=temperature,
        max_tokens=max_tokens,
    )


def ask_json(system: str, user: str, temperature: float = 0.3) -> dict:
    """Calls the API and parses the JSON response — strips markdown fences if any."""
    full_system = (
        system
        + "\n\nCRITICAL: Respond with a valid JSON object ONLY."
          " No markdown, no code fences, no explanation outside the JSON."
    )
    text = nvidia_call(
        [{"role": "system", "content": full_system}, {"role": "user", "content": user}],
        temperature=temperature,
        max_tokens=500,
    )
    # Strip accidental ```json ... ``` wrappers
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:])
    if text.endswith("```"):
        text = text.rsplit("```", 1)[0]
    return json.loads(text.strip())


SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")
SENDGRID_FROM_EMAIL = os.getenv("SENDGRID_FROM_EMAIL", "")
SENDGRID_TO_EMAIL = os.getenv("SENDGRID_TO_EMAIL", "")


def send_email_notification(subject: str, body: str) -> None:
    if not SENDGRID_API_KEY or not SENDGRID_FROM_EMAIL or not SENDGRID_TO_EMAIL:
        return

    response = requests.post(
        "https://api.sendgrid.com/v3/mail/send",
        headers={
            "Authorization": f"Bearer {SENDGRID_API_KEY}",
            "Content-Type": "application/json",
        },
        json={
            "personalizations": [{"to": [{"email": SENDGRID_TO_EMAIL}]}],
            "from": {"email": SENDGRID_FROM_EMAIL},
            "subject": subject,
            "content": [{"type": "text/plain", "value": body}],
        },
        timeout=20,
    )
    response.raise_for_status()


def parse_json_list(value):
    if isinstance(value, list):
        return value
    if not value:
        return []
    try:
        data = json.loads(value)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def parse_bearer_token(authorization: str = "") -> str:
    if not authorization:
        return ""
    parts = authorization.split(" ", 1)
    return parts[1].strip() if len(parts) == 2 and parts[0].lower() == "bearer" else authorization.strip()


def require_admin_token(authorization: str = "") -> str:
    token = parse_bearer_token(authorization)
    if not verify_token(token):
        raise HTTPException(status_code=401, detail="Unauthorized")
    return token


def serialize_learning_topic(topic_row):
    topic = dict(topic_row)
    topic["links"] = parse_json_list(topic.pop("links_json", "[]"))
    topic["tags"] = parse_json_list(topic.pop("tags_json", "[]"))
    topic["is_news"] = bool(topic.get("is_news"))
    return topic


def serialize_learning_track(track_row, topics):
    track = dict(track_row)
    track["tags"] = parse_json_list(track.pop("tags_json", "[]"))
    track["is_news"] = bool(track.get("is_news"))
    track["topics"] = topics
    return track


GAME_SYSTEM = (
    "You are an elite 'Guess Who / 20 Questions' deduction engine for famous real people. "
    "Your job is to maximize accuracy, not to guess early. "
    "Ask globally discriminating questions first, then narrow by domain, era, geography, and signature achievements. "
    "Treat 'Probably' and 'Probably not' as soft evidence, not certainty. "
    "Only make a final guess when the clue set is genuinely specific enough. "
    "You reason step by step internally before answering. "
    "Respond with valid JSON ONLY — no markdown, no fences, no extra text."
)

# ─── SCHEMAS ───────────────────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    message: str
    history: List[dict] = []

class HireRequest(BaseModel):
    question: str
    history: List[dict] = []
    company: Optional[str] = ""

class GameStartRequest(BaseModel):
    player_name: str

class GameAnswerRequest(BaseModel):
    session_id: str
    question_number: int
    clue: str
    all_clues: List[str]

game_sessions: dict = {}


# ─── CHAT ──────────────────────────────────────────────────────────────────────
@app.post("/chat")
async def chat(req: ChatRequest):
    try:
        messages = [{"role": "system", "content": CLONE_SYSTEM_PROMPT}]
        for h in req.history[-10:]:
            messages.append({
                "role": "user" if h["role"] == "user" else "assistant",
                "content": h["content"],
            })
        messages.append({"role": "user", "content": req.message})
        reply = nvidia_call(messages, temperature=0.85, max_tokens=350)
        return {"reply": reply}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── HIRE ME ───────────────────────────────────────────────────────────────────
@app.post("/hire")
async def hire(req: HireRequest):
    """Answers recruiter / HR questions AS the candidate."""
    try:
        system = HIRE_SYSTEM_PROMPT
        if req.company:
            system += f"\n\nNOTE: The recruiter is from '{req.company}'. Tailor your answer accordingly."

        messages = [{"role": "system", "content": system}]
        for h in req.history[-8:]:
            messages.append({
                "role": "user" if h["role"] == "user" else "assistant",
                "content": h["content"],
            })
        messages.append({"role": "user", "content": req.question})
        reply = nvidia_call(messages, temperature=0.6, max_tokens=500)
        return {"reply": reply}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── IDLE NUDGE ────────────────────────────────────────────────────────────────
@app.get("/idle-nudge")
async def idle_nudge():
    try:
        user = """The person you're chatting with has gone completely silent.
Send ONE short funny casual nudge. Pick a random style:
- Malayalam slang: "eda enthu patti? 😴" or "entha moluse jadayan aayo?"
- Dramatic: "HELLO?? IS THIS THING ON?? 📡"
- Petty/sad: "...fine I'll just talk to the void" or "guess I'll go touch grass 🌱"
- Playful: "bro did you actually fall asleep on me 💀"
Return ONLY the nudge message — no quotes, nothing else."""
        msg = ask(CLONE_SYSTEM_PROMPT, user, temperature=1.0, max_tokens=60)
        return {"message": msg}
    except:
        opts = [
            "eda, enthu patti? 👀",
            "entha moluse jadayan aayo? 😴",
            "hello?? earth to human 📡",
            "bro did you die 💀",
            "...okay I'll just stare at the void then",
            "fine. I'll talk to myself. I'm great company anyway.",
        ]
        return {"message": random.choice(opts)}


# ─── QUAKE REACTION ────────────────────────────────────────────────────────────
@app.get("/quake-react")
async def quake_react():
    try:
        user = """ONE ultra-short (max 8 words) caps-heavy dramatic funny message
an AI screams when the screen shakes violently.
Examples: "AAAAAA WE ARE ALL DYING 😱" or "ENTHO SAMBHAVIKKUNU 😱 HELP"
Return ONLY the message — no quotes."""
        msg = ask("You are a dramatic funny AI.", user, temperature=1.1, max_tokens=40)
        return {"message": msg}
    except:
        opts = [
            "AAAAAA WE ARE ALL GONNA DIE 😱",
            "SYSTEM MELTDOWN EVERYBODY RUN 🚨",
            "ENTHO SAMBHAVIKKUNU 😱 HELP",
            "I AM TOO YOUNG TO CRASH 💀",
            "WE ARE ENDING BRO 😱😱😱",
        ]
        return {"message": random.choice(opts)}


# ─── GAME: START ───────────────────────────────────────────────────────────────
@app.post("/game/start")
async def game_start(req: GameStartRequest):
    sid = str(uuid.uuid4())[:8]
    game_sessions[sid] = {"player": req.player_name, "q_count": 0}
    try:
        data = ask_json(
            GAME_SYSTEM,
            f"""Start a Guess Who game with player: {req.player_name}
They are thinking of a famous person (any nationality, living or historical, real person).
You can use as many questions as needed, but keep each question highly informative.

The BEST first question should split the space of famous people into large useful groups.
Classic opener: "Is this person currently alive?"

Return a JSON object with exactly these two keys:
{{
  "greeting": "short fun greeting to {req.player_name}, can use Malayalam slang like eda/machi",
  "first_question": "your single best binary-search yes/no opening question"
}}""",
            temperature=0.6,
        )
        return {"session_id": sid, "first_question": data["first_question"], "greeting": data["greeting"]}
    except:
        return {
            "session_id": sid,
            "first_question": "Is this person currently alive?",
            "greeting": f"Okay {req.player_name} da, I am going for accuracy this time — you're not escaping me 😏",
        }


# ─── GAME: ANSWER ──────────────────────────────────────────────────────────────
@app.post("/game/answer")
async def game_answer(req: GameAnswerRequest):
    clues = "\n".join([f"  Q{i+1}: {c}" for i, c in enumerate(req.all_clues)])
    q_num = req.question_number

    # ── PHASE 1 (Q1–Q8): Broad category questions ────────────────────
    # Just ask — too early to guess

    # ── PHASE 2 (Q9+): Check confidence after every answer ───────────
    if q_num >= 12:
        try:
            check = ask_json(
                GAME_SYSTEM,
                f"""You are playing Guess Who. Clues collected so far ({q_num} questions):
{clues}

Evaluate your confidence level carefully.
Consider ALL famous people worldwide: actors, athletes, politicians, musicians,
scientists, business leaders, historical figures, YouTubers, artists.
Do NOT guess just because one candidate feels plausible.
Only set can_guess=true if the current clues strongly rule out close alternatives.

Return a JSON object:
{{
  "can_guess": true or false,
  "guess": "Full Name (or null if not confident)",
  "confidence_pct": number from 0 to 100,
  "top_candidates": ["Name1", "Name2", "Name3"]
}}""",
                temperature=0.15,
            )
            conf = int(check.get("confidence_pct", 0))

            threshold = 96 if q_num <= 16 else (92 if q_num <= 22 else (86 if q_num <= 30 else 78))

            if check.get("can_guess") and conf >= threshold:
                reason = ask(
                    GAME_SYSTEM,
                    f"You guessed '{check['guess']}' in Guess Who from these clues:\n{clues}\n"
                    "Write ONE fun casual sentence explaining your deduction logic. No JSON.",
                    temperature=0.9,
                    max_tokens=100,
                )
                return {
                    "ai_guess": check["guess"],
                    "reasoning": reason,
                    "confidence": "high" if conf >= 85 else "medium",
                    "is_final": True,
                }
        except:
            pass  # not confident, continue asking

    # ── PHASE 3 (Q25): Force final guess ─────────────────────────────
    if q_num >= 35:
        try:
            data = ask_json(
                GAME_SYSTEM,
                f"""You have reached the loop safety stop. Here are ALL the clues:
{clues}

You MUST make your absolute best guess now — no more questions allowed.
Think carefully: which single famous person matches ALL these clues perfectly?
Consider every possible category of famous person globally.

Return a JSON object:
{{
  "guess": "Full Name",
  "confidence": "high" or "medium" or "low",
  "reasoning": "one fun casual sentence explaining your final deduction"
}}""",
                temperature=0.1,
            )
            return {
                "ai_guess": data["guess"],
                "reasoning": data.get("reasoning", ""),
                "confidence": data["confidence"],
                "is_final": True,
            }
        except:
            return {
                "ai_guess": "I genuinely have no idea 😅 You completely stumped me!",
                "reasoning": "",
                "confidence": "low",
                "is_final": True,
            }

    # ── ASK NEXT SMART QUESTION ───────────────────────────────────────
    # Build a phase-aware prompt so questions get more specific over time
    if q_num < 8:
        strategy = """You are in the BROAD PHASE (questions 1-8).
Ask questions that split the entire space of famous people in half:
alive/dead → gender → nationality/continent → field (entertainment/politics/sports/science/business) → era → sub-field"""
    elif q_num < 18:
        strategy = """You are in the NARROWING PHASE (questions 9-16).
The broad category is established. Now narrow down:
→ specific country within region → specific decade they became famous
→ specific sub-field (e.g., actor vs musician vs athlete)
→ major awards or achievements → whether they are still active"""
    else:
        strategy = """You are in the PRECISION PHASE (later questions).
You should be close. Ask very specific questions:
→ first letter or length of name → specific role/film/sport/achievement they're most known for
→ physical characteristics if relevant → specific year/event
→ any unique distinguishing fact"""

    try:
        question = ask(
            GAME_SYSTEM,
            f"""You are guessing a famous person. {strategy}

Clues collected so far:
{clues}

Rules for this question:
- Must be a YES/NO question only
- NEVER repeat or rephrase something already answered in the clues above
- Prefer questions with high information gain over flashy questions
- Make it count — this is question #{q_num + 1}

Return ONLY the question text. No explanation, no numbering, no extra words.""",
            temperature=0.4,
            max_tokens=70,
        )
        question = question.strip().strip('"').strip("'")
        if not question.endswith("?"):
            question += "?"
        return {"next_question": question, "is_final": False}

    except:
        # Phase-aware fallbacks
        early = [
            "Is this person currently alive?",
            "Is this person male?",
            "Is this person primarily known for entertainment?",
            "Is this person from Asia?",
            "Did this person become famous after the year 2000?",
            "Is this person known internationally outside their home country?",
            "Is this person an athlete or sports personality?",
            "Is this person a politician or political leader?",
        ]
        mid = [
            "Has this person won a major international award?",
            "Is this person from India?",
            "Is this person primarily known for a single film, song, or achievement?",
            "Is this person under 50 years old?",
            "Is this person active on social media with millions of followers?",
            "Did this person become famous primarily in the 1990s or earlier?",
            "Is this person known by a single name or short nickname?",
            "Is this person associated with a specific sport?",
        ]
        late = [
            "Is this person's first name a common name?",
            "Does this person's name start with a letter in the first half of the alphabet (A-M)?",
            "Is this person primarily known for one specific iconic thing?",
            "Is this person considered a legend or GOAT in their field?",
            "Has this person appeared in a major Hollywood or Bollywood production?",
            "Is this person known for a specific physical appearance or signature style?",
        ]
        pool = early if q_num < 8 else (mid if q_num < 16 else late)
        idx = q_num % len(pool)
        return {"next_question": pool[idx], "is_final": False}


# ─── JD MATCH ──────────────────────────────────────────────────────────────────
class JDMatchRequest(BaseModel):
    jd_text: str

@app.post("/jd-match")
async def jd_match(req: JDMatchRequest):
    """AI-powered JD match analysis — returns score, matched skills, gaps, summary."""
    try:
        from profile import MY_KNOWLEDGE
        system = """You are an expert technical recruiter and career coach.
Analyse how well a candidate's profile matches a job description.
Be accurate, specific, and helpful. Respond with valid JSON ONLY."""

        user = f"""Here is the candidate's full profile:
{MY_KNOWLEDGE}

Here is the job description to match against:
{req.jd_text}

Analyse the match thoroughly and return a JSON object with exactly these keys:
{{
  "score": a number 0-100 representing overall match percentage,
  "verdict": one of "Strong Match", "Good Match", "Partial Match", "Weak Match",
  "summary": "2-3 sentence honest summary of how well the candidate fits this role",
  "matched_skills": ["list", "of", "skills", "candidate", "has", "that", "JD", "needs"],
  "missing_skills": ["list", "of", "skills", "in", "JD", "candidate", "lacks"],
  "relevant_projects": ["project names that are most relevant to this role"],
  "strengths_for_role": ["2-3 specific strengths that make candidate good for this role"],
  "improvement_areas": ["1-2 honest areas to improve for this role"],
  "interview_tips": ["2 specific tips for interviewing for this exact role"]
}}"""

        data = ask_json(system, user, temperature=0.2)
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── CONTACT FORM ──────────────────────────────────────────────────────────────
class ContactRequest(BaseModel):
    name: str
    email: str
    company: str = ""
    message: str

CONTACTS_EXCEL = os.path.join(os.path.dirname(__file__), "contacts.xlsx")

def save_contact_to_excel(name: str, email: str, company: str, message: str, ai_reply: str):
    """Appends a contact submission to contacts.xlsx — creates the file if it doesn't exist."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(CONTACTS_EXCEL):
        wb = load_workbook(CONTACTS_EXCEL)
        ws = wb.active
    else:
        # ── Create fresh workbook with styled header ──────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        headers = ["#", "Date & Time", "Name", "Email", "Company", "Message", "AI Reply", "Status"]
        col_widths = [5, 20, 20, 28, 22, 50, 50, 12]

        # Header style
        header_fill   = PatternFill("solid", fgColor="3B1F8C")   # deep violet
        header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin_border   = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"   # freeze header row

    # ── Append new row ────────────────────────────────────────────────
    row_num    = ws.max_row + 1
    entry_num  = ws.max_row   # row 1 = header, so row 2 = entry #1

    row_fill  = PatternFill("solid", fgColor="F4F0FF") if entry_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    left_al   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    center_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_bdr  = Border(
        left=Side(style="thin",   color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
        top=Side(style="thin",    color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    row_data = [entry_num, now, name, email, company or "—", message, ai_reply, "New"]
    alignments = [center_al, center_al, left_al, left_al, left_al, left_al, left_al, center_al]

    for col_idx, (value, align) in enumerate(zip(row_data, alignments), start=1):
        cell            = ws.cell(row=row_num, column=col_idx, value=value)
        cell.alignment  = align
        cell.fill       = row_fill
        cell.border     = thin_bdr
        cell.font       = Font(name="Calibri", size=10)

    # Colour-code the Status cell
    status_cell       = ws.cell(row=row_num, column=8)
    status_cell.font  = Font(name="Calibri", bold=True, color="1A6E2E", size=10)
    status_cell.fill  = PatternFill("solid", fgColor="D4EDDA")

    # Auto-fit row height roughly based on message length
    max_chars = max(len(message), len(ai_reply))
    ws.row_dimensions[row_num].height = max(20, min(80, max_chars // 3))

    wb.save(CONTACTS_EXCEL)
    print(f"✅ Saved to contacts.xlsx — row {entry_num}")


@app.post("/contact")
async def contact(req: ContactRequest):
    """Logs contact form submission to Excel and returns AI-generated acknowledgement."""
    try:
        system = CLONE_SYSTEM_PROMPT
        user = f"""Someone just sent you a message through your AI Clone portfolio.

From: {req.name}
Email: {req.email}
Company: {req.company or 'Not specified'}
Message: {req.message}

Write a SHORT, warm, genuine acknowledgement reply (3-4 sentences max) in your personality.
Be professional but casual. Thank them, address their message briefly, and say you'll follow up personally.
Return ONLY the reply text — no JSON, no quotes around it."""

        reply = ask(system, user, temperature=0.75, max_tokens=200)

        # ── Log to console ────────────────────────────────────────────
        print(f"\n📬 NEW CONTACT from {req.name} ({req.email})"
              f"\nCompany : {req.company or '—'}"
              f"\nMessage : {req.message}\n")

        # ── Save to Excel (non-blocking — errors won't break the API) ─
        try:
            save_contact_to_excel(req.name, req.email, req.company, req.message, reply)
        except Exception as excel_err:
            print(f"⚠️  Excel save failed: {excel_err}")

        try:
            send_email_notification(
                f"New AI Clone contact from {req.name}",
                (
                    f"New contact form submission\n\n"
                    f"Name: {req.name}\n"
                    f"Email: {req.email}\n"
                    f"Company: {req.company or '-'}\n\n"
                    f"Message:\n{req.message}\n\n"
                    f"AI reply:\n{reply}\n"
                ),
            )
        except Exception as email_err:
            print(f"Contact email failed: {email_err}")

        return {"success": True, "reply": reply, "message": "Message received successfully!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/contacts/download")
async def download_contacts():
    """Download the contacts Excel file directly from the browser."""
    from fastapi.responses import FileResponse
    if not os.path.exists(CONTACTS_EXCEL):
        raise HTTPException(status_code=404, detail="No contacts yet — file will be created on first submission.")
    return FileResponse(
        path=CONTACTS_EXCEL,
        filename="ai_clone_contacts.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/contacts/count")
async def contacts_count():
    """Returns how many contact submissions exist."""
    if not os.path.exists(CONTACTS_EXCEL):
        return {"count": 0}
    from openpyxl import load_workbook
    wb = load_workbook(CONTACTS_EXCEL, read_only=True)
    count = wb.active.max_row - 1   # subtract header row
    return {"count": max(count, 0)}



# ─── REFERRALS ─────────────────────────────────────────────────────────────────

@app.post("/referrals/add")
async def add_referral(
    name:     str           = Form(...),
    quote:    str           = Form(...),
    email:    str           = Form(""),
    role:     str           = Form(""),
    company:  str           = Form(""),
    relation: str           = Form(""),
    photo:    Optional[UploadFile] = File(None),
):
    """Add a referral/testimonial with optional photo upload."""
    photo_path = None

    if photo and photo.filename:
        ext        = os.path.splitext(photo.filename)[1].lower()
        allowed    = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
        if ext not in allowed:
            raise HTTPException(400, "Photo must be jpg/png/gif/webp")
        fname      = f"{uuid.uuid4().hex}{ext}"
        dest       = os.path.join(UPLOADS_DIR, fname)
        with open(dest, "wb") as f:
            shutil.copyfileobj(photo.file, f)
        photo_path = f"/uploads/{fname}"

    conn = get_db()
    try:
        conn.execute(
            """INSERT INTO referrals (name, email, role, company, quote, photo_path, relation)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (name, email, role, company, quote, photo_path, relation),
        )
        conn.commit()
        return {"success": True, "message": "Referral added!", "photo_url": photo_path}
    finally:
        conn.close()


@app.get("/referrals")
async def get_referrals():
    """Return all referrals ordered newest first."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM referrals ORDER BY id DESC"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.delete("/referrals/{ref_id}")
async def delete_referral(ref_id: int):
    """Delete a referral and its photo file."""
    conn = get_db()
    try:
        row = conn.execute("SELECT photo_path FROM referrals WHERE id=?", (ref_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Referral not found")
        # Delete photo from disk if it exists
        if row["photo_path"]:
            full = os.path.join(BASE_DIR, row["photo_path"].lstrip("/"))
            if os.path.exists(full):
                os.remove(full)
        conn.execute("DELETE FROM referrals WHERE id=?", (ref_id,))
        conn.commit()
        return {"success": True}
    finally:
        conn.close()


# ─── FEEDBACK ──────────────────────────────────────────────────────────────────

class FeedbackRequest(BaseModel):
    message: str
    rating:  int = 5


class LearningTrackPayload(BaseModel):
    title: str
    description: str = ""
    kind: str = "tree"
    priority: int = 0
    color: str = ""
    tags: List[str] = []
    is_news: bool = False


class LearningTopicLink(BaseModel):
    label: str
    url: str


class LearningTopicPayload(BaseModel):
    track_id: int
    parent_id: Optional[int] = None
    title: str
    description: str = ""
    priority: int = 0
    links: List[LearningTopicLink] = []
    tags: List[str] = []
    is_news: bool = False


class LearningProgressPayload(BaseModel):
    topic_id: int
    fingerprint: str
    status: str


class LearningCommentPayload(BaseModel):
    name: str = ""
    body: str
    parent_id: Optional[int] = None
    fingerprint: str = ""


@app.post("/feedback/add")
async def add_feedback(req: FeedbackRequest):
    """Save anonymous feedback with optional 1-5 star rating."""
    if not req.message.strip():
        raise HTTPException(400, "Feedback message cannot be empty")
    rating = max(1, min(5, req.rating))
    conn   = get_db()
    try:
        conn.execute(
            "INSERT INTO feedback (message, rating) VALUES (?, ?)",
            (req.message.strip(), rating),
        )
        conn.commit()

        # AI generates a quick thanks message
        try:
            thanks = ask(
                CLONE_SYSTEM_PROMPT,
                f"Someone left you {rating}-star feedback on your portfolio: \"{req.message[:200]}\"\n"
                "Write ONE short casual genuine thank-you (2 sentences max). No JSON.",
                temperature=0.8, max_tokens=80,
            )
        except:
            thanks = "Thanks so much for the feedback! Really appreciate you taking the time 🙏"

        try:
            send_email_notification(
                f"New AI Clone feedback ({rating}/5)",
                (
                    f"New feedback received\n\n"
                    f"Rating: {rating}/5\n\n"
                    f"Message:\n{req.message.strip()}\n\n"
                    f"AI reply:\n{thanks}\n"
                ),
            )
        except Exception as email_err:
            print(f"Feedback email failed: {email_err}")

        return {"success": True, "reply": thanks}
    finally:
        conn.close()


@app.get("/feedback")
async def get_feedback():
    """Return all feedback ordered newest first."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM feedback ORDER BY id DESC"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/feedback/stats")
async def feedback_stats():
    """Returns count and average rating of all feedback."""
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT COUNT(*) as count, ROUND(AVG(rating),1) as avg_rating FROM feedback"
        ).fetchone()
        return {"count": row["count"], "avg_rating": row["avg_rating"] or 0}
    finally:
        conn.close()

# ─── VOICE TTS ─────────────────────────────────────────────────────────────────
@app.get("/learning/tracks")
def get_learning_tracks(kind: Optional[str] = None):
    conn = get_db()
    try:
        query = "SELECT * FROM learning_tracks"
        params = []
        if kind:
            query += " WHERE kind = ?"
            params.append(kind)
        query += " ORDER BY priority DESC, updated_at DESC, id DESC"
        track_rows = conn.execute(query, params).fetchall()
        track_ids = [row["id"] for row in track_rows]
        topics_by_track = {track_id: [] for track_id in track_ids}
        if track_ids:
            placeholders = ",".join("?" for _ in track_ids)
            topic_rows = conn.execute(
                f"SELECT * FROM learning_topics WHERE track_id IN ({placeholders}) ORDER BY priority DESC, id ASC",
                track_ids,
            ).fetchall()
            for row in topic_rows:
                topics_by_track[row["track_id"]].append(serialize_learning_topic(row))
        return [serialize_learning_track(row, topics_by_track.get(row["id"], [])) for row in track_rows]
    finally:
        conn.close()


@app.post("/learning/tracks")
def create_learning_track(payload: LearningTrackPayload, authorization: str = Header(default="")):
    require_admin_token(authorization)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO learning_tracks (title, description, kind, priority, color, tags_json, is_news)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.title.strip(),
                payload.description.strip(),
                payload.kind.strip() or "tree",
                payload.priority,
                payload.color.strip(),
                json.dumps(payload.tags),
                1 if payload.is_news else 0,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM learning_tracks WHERE id = ?", (cur.lastrowid,)).fetchone()
        return serialize_learning_track(row, [])
    finally:
        conn.close()


@app.put("/learning/tracks/{track_id}")
def update_learning_track(track_id: int, payload: LearningTrackPayload, authorization: str = Header(default="")):
    require_admin_token(authorization)
    conn = get_db()
    try:
        conn.execute(
            """
            UPDATE learning_tracks
            SET title = ?, description = ?, kind = ?, priority = ?, color = ?, tags_json = ?, is_news = ?, updated_at = datetime('now','localtime')
            WHERE id = ?
            """,
            (
                payload.title.strip(),
                payload.description.strip(),
                payload.kind.strip() or "tree",
                payload.priority,
                payload.color.strip(),
                json.dumps(payload.tags),
                1 if payload.is_news else 0,
                track_id,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM learning_tracks WHERE id = ?", (track_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Track not found")
        topic_rows = conn.execute("SELECT * FROM learning_topics WHERE track_id = ? ORDER BY priority DESC, id ASC", (track_id,)).fetchall()
        return serialize_learning_track(row, [serialize_learning_topic(topic) for topic in topic_rows])
    finally:
        conn.close()


@app.delete("/learning/tracks/{track_id}")
def delete_learning_track(track_id: int, authorization: str = Header(default="")):
    require_admin_token(authorization)
    conn = get_db()
    try:
        conn.execute("DELETE FROM learning_topics WHERE track_id = ?", (track_id,))
        conn.execute("DELETE FROM learning_tracks WHERE id = ?", (track_id,))
        conn.commit()
        return {"deleted": True}
    finally:
        conn.close()


@app.post("/learning/topics")
def create_learning_topic(payload: LearningTopicPayload, authorization: str = Header(default="")):
    require_admin_token(authorization)
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO learning_topics (track_id, parent_id, title, description, priority, links_json, tags_json, is_news)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload.track_id,
                payload.parent_id,
                payload.title.strip(),
                payload.description.strip(),
                payload.priority,
                json.dumps([link.dict() for link in payload.links]),
                json.dumps(payload.tags),
                1 if payload.is_news else 0,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM learning_topics WHERE id = ?", (cur.lastrowid,)).fetchone()
        return serialize_learning_topic(row)
    finally:
        conn.close()


@app.put("/learning/topics/{topic_id}")
def update_learning_topic(topic_id: int, payload: LearningTopicPayload, authorization: str = Header(default="")):
    require_admin_token(authorization)
    conn = get_db()
    try:
        conn.execute(
            """
            UPDATE learning_topics
            SET track_id = ?, parent_id = ?, title = ?, description = ?, priority = ?, links_json = ?, tags_json = ?, is_news = ?, updated_at = datetime('now','localtime')
            WHERE id = ?
            """,
            (
                payload.track_id,
                payload.parent_id,
                payload.title.strip(),
                payload.description.strip(),
                payload.priority,
                json.dumps([link.dict() for link in payload.links]),
                json.dumps(payload.tags),
                1 if payload.is_news else 0,
                topic_id,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM learning_topics WHERE id = ?", (topic_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Topic not found")
        return serialize_learning_topic(row)
    finally:
        conn.close()


@app.delete("/learning/topics/{topic_id}")
def delete_learning_topic(topic_id: int, authorization: str = Header(default="")):
    require_admin_token(authorization)
    conn = get_db()
    try:
        conn.execute("DELETE FROM learning_topics WHERE id = ? OR parent_id = ?", (topic_id, topic_id))
        conn.commit()
        return {"deleted": True}
    finally:
        conn.close()


@app.get("/learning/tracks/{track_id}/progress")
def get_learning_progress(track_id: int, fingerprint: str):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT topic_id, status FROM learning_progress WHERE track_id = ? AND fingerprint = ?",
            (track_id, fingerprint),
        ).fetchall()
        return {str(row["topic_id"]): row["status"] for row in rows}
    finally:
        conn.close()


@app.post("/learning/tracks/{track_id}/progress")
def set_learning_progress(track_id: int, payload: LearningProgressPayload):
    status = (payload.status or "").strip().lower()
    if status not in {"pending", "in_progress", "completed"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    if not payload.fingerprint.strip():
        raise HTTPException(status_code=400, detail="Fingerprint is required")

    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO learning_progress (track_id, topic_id, fingerprint, status, updated_at)
            VALUES (?, ?, ?, ?, datetime('now','localtime'))
            ON CONFLICT(topic_id, fingerprint) DO UPDATE SET
                status = excluded.status,
                updated_at = datetime('now','localtime')
            """,
            (track_id, payload.topic_id, payload.fingerprint.strip(), status),
        )
        conn.commit()
        return {"success": True, "status": status}
    finally:
        conn.close()


@app.get("/learning/tracks/{track_id}/comments")
def get_learning_comments(track_id: int):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM learning_comments WHERE track_id = ? ORDER BY created_at ASC, id ASC",
            (track_id,),
        ).fetchall()
        return [{**dict(row), "is_author": bool(row["is_author"])} for row in rows]
    finally:
        conn.close()


@app.post("/learning/tracks/{track_id}/comments")
def add_learning_comment(track_id: int, payload: LearningCommentPayload, authorization: str = Header(default="")):
    if not payload.body.strip():
        raise HTTPException(status_code=400, detail="Comment cannot be empty")

    token = parse_bearer_token(authorization)
    author_mode = bool(token and verify_token(token))
    name = (payload.name or "").strip()
    fingerprint = (payload.fingerprint or "").strip()

    if author_mode:
        name = name or "Course Author"
    else:
        if not name:
            raise HTTPException(status_code=400, detail="Name is required")
        if not fingerprint:
            raise HTTPException(status_code=400, detail="Fingerprint is required")

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO learning_comments (track_id, parent_id, name, body, fingerprint, is_author)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (track_id, payload.parent_id, name, payload.body.strip(), fingerprint, 1 if author_mode else 0),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM learning_comments WHERE id = ?", (cur.lastrowid,)).fetchone()
        return {**dict(row), "is_author": bool(row["is_author"])}
    finally:
        conn.close()


ELEVENLABS_API_KEY = os.getenv("ELEVENLABS_API_KEY", "")
ELEVENLABS_VOICE_ID = os.getenv("ELEVENLABS_VOICE_ID", "")

class VoiceRequest(BaseModel):
    text: str

@app.post("/voice/speak")
async def text_to_speech(req: VoiceRequest):
    """Convert text to speech using ElevenLabs voice clone."""
    if not ELEVENLABS_API_KEY or not ELEVENLABS_VOICE_ID:
        raise HTTPException(
            status_code=503,
            detail="ElevenLabs not configured. Add ELEVENLABS_API_KEY and ELEVENLABS_VOICE_ID to your .env file."
        )

    text = req.text.strip()[:600]  # cap to save credits
    if not text:
        raise HTTPException(status_code=400, detail="Text is empty")

    try:
        response = requests.post(
            f"https://api.elevenlabs.io/v1/text-to-speech/{ELEVENLABS_VOICE_ID}",
            headers={
                "xi-api-key": ELEVENLABS_API_KEY,
                "Content-Type": "application/json",
                "Accept": "audio/mpeg",
            },
            json={
                "text": text,
                "model_id": "eleven_monolingual_v1",
                "voice_settings": {
                    "stability": 0.5,
                    "similarity_boost": 0.75,
                    "style": 0.3,
                    "use_speaker_boost": True
                }
            },
            timeout=15
        )

        if response.status_code != 200:
            raise HTTPException(status_code=response.status_code, detail="ElevenLabs API error")

        import base64
        audio_base64 = base64.b64encode(response.content).decode("utf-8")
        return {"audio": audio_base64, "format": "audio/mpeg"}

    except requests.exceptions.Timeout:
        raise HTTPException(status_code=504, detail="ElevenLabs request timed out")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ─── BLOG ADMIN AUTH ───────────────────────────────────────────────────────────
class AdminLoginRequest(BaseModel):
    username: str
    password: str

def verify_token(token: str) -> bool:
    """Simple token verification — token is sha256(username+password+salt)."""
    if not token:
        return False
    conn = get_db()
    try:
        row = conn.execute("SELECT 1 FROM blog_admins WHERE password = ?", (token,)).fetchone()
        return row is not None
    finally:
        conn.close()

@app.post("/blog-admin/login")
def blog_admin_login(req: AdminLoginRequest):
    conn = get_db()
    try:
        hashed = hashlib.sha256(req.password.encode()).hexdigest()
        row = conn.execute(
            "SELECT id FROM blog_admins WHERE username = ? AND password = ?",
            (req.username, hashed)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=401, detail="Invalid username or password")
        # Token is just the hashed password — simple but works for single admin
        return {"token": hashed, "message": "Login successful"}
    finally:
        conn.close()

@app.post("/blog-admin/change-password")
def change_password(req: dict):
    token     = req.get("token", "")
    new_pass  = req.get("new_password", "")
    if not verify_token(token):
        raise HTTPException(status_code=401, detail="Unauthorized")
    if len(new_pass) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    new_hash = hashlib.sha256(new_pass.encode()).hexdigest()
    conn = get_db()
    try:
        conn.execute("UPDATE blog_admins SET password = ? WHERE password = ?", (new_hash, token))
        conn.commit()
        return {"token": new_hash, "message": "Password changed"}
    finally:
        conn.close()


# ─── BLOG CRUD ─────────────────────────────────────────────────────────────────
class BlogCreate(BaseModel):
    title:       str
    description: str
    category:    str = "Tech"
    image_url:   str = ""

class CommentCreate(BaseModel):
    name: str = "Anonymous"
    body: str

@app.get("/blogs")
def get_blogs():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT b.*, 
                   (SELECT COUNT(*) FROM blog_comments c WHERE c.blog_id = b.id) as comment_count
            FROM blogs b ORDER BY b.created_at DESC
        """).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

@app.get("/blogs/{blog_id}")
def get_blog(blog_id: int):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM blogs WHERE id = ?", (blog_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Blog not found")
        return dict(row)
    finally:
        conn.close()

@app.post("/blogs")
def create_blog(blog: BlogCreate, authorization: str = ""):
    # Extract token from Authorization header via Request
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO blogs (title, description, category, image_url) VALUES (?, ?, ?, ?)",
            (blog.title.strip(), blog.description.strip(), blog.category, blog.image_url)
        )
        conn.commit()
        blog_id = cur.lastrowid
        row = conn.execute("SELECT * FROM blogs WHERE id = ?", (blog_id,)).fetchone()
        return dict(row)
    finally:
        conn.close()

@app.delete("/blogs/{blog_id}")
def delete_blog(blog_id: int):
    conn = get_db()
    try:
        conn.execute("DELETE FROM blogs WHERE id = ?", (blog_id,))
        conn.execute("DELETE FROM blog_comments WHERE blog_id = ?", (blog_id,))
        conn.execute("DELETE FROM blog_likes WHERE blog_id = ?", (blog_id,))
        conn.commit()
        return {"deleted": True}
    finally:
        conn.close()

@app.post("/blogs/upload-image")
async def upload_blog_image(file: UploadFile = File(...)):
    ext      = os.path.splitext(file.filename)[1].lower()
    allowed  = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="Image files only")
    fname    = f"blog_{uuid.uuid4().hex}{ext}"
    fpath    = os.path.join(UPLOADS_DIR, fname)
    with open(fpath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"url": f"/uploads/{fname}"}

# ─── BLOG LIKES ────────────────────────────────────────────────────────────────
@app.post("/blogs/{blog_id}/like")
def like_blog(blog_id: int, request_data: dict = {}):
    # Use a fingerprint from request — in production you'd use IP
    # Here we accept an optional fingerprint from frontend or generate one
    fingerprint = request_data.get("fingerprint", str(uuid.uuid4()))
    conn = get_db()
    try:
        # Try to insert unique like
        try:
            conn.execute(
                "INSERT INTO blog_likes (blog_id, fingerprint) VALUES (?, ?)",
                (blog_id, fingerprint)
            )
            conn.execute("UPDATE blogs SET likes = likes + 1 WHERE id = ?", (blog_id,))
            conn.commit()
            row = conn.execute("SELECT likes FROM blogs WHERE id = ?", (blog_id,)).fetchone()
            return {"likes": row["likes"], "liked": True}
        except sqlite3.IntegrityError:
            # Already liked
            row = conn.execute("SELECT likes FROM blogs WHERE id = ?", (blog_id,)).fetchone()
            return {"likes": row["likes"] if row else 0, "liked": False, "message": "Already liked"}
    finally:
        conn.close()

# ─── BLOG COMMENTS ─────────────────────────────────────────────────────────────
@app.get("/blogs/{blog_id}/comments")
def get_comments(blog_id: int):
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM blog_comments WHERE blog_id = ? ORDER BY created_at ASC",
            (blog_id,)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

@app.post("/blogs/{blog_id}/comments")
def add_comment(blog_id: int, comment: CommentCreate):
    if not comment.body.strip():
        raise HTTPException(status_code=400, detail="Comment cannot be empty")
    conn = get_db()
    try:
        # Check blog exists
        if not conn.execute("SELECT 1 FROM blogs WHERE id = ?", (blog_id,)).fetchone():
            raise HTTPException(status_code=404, detail="Blog not found")
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO blog_comments (blog_id, name, body) VALUES (?, ?, ?)",
            (blog_id, comment.name.strip() or "Anonymous", comment.body.strip())
        )
        conn.commit()
        row = conn.execute("SELECT * FROM blog_comments WHERE id = ?", (cur.lastrowid,)).fetchone()
        return dict(row)
    finally:
        conn.close()

# ─── ROOT ──────────────────────────────────────────────────────────────────────
@app.get("/")
def root():
    return {
        "status": "AI Clone running 🚀",
        "model": MODEL,
        "endpoints": ["/chat", "/hire", "/jd-match", "/contact", "/referrals", "/referrals/add", "/feedback/add", "/feedback", "/feedback/stats", "/game/start", "/game/answer", "/idle-nudge", "/quake-react", "/blogs", "/blog-admin/login", "/learning/tracks"],
        "docs": "/docs",
    }
